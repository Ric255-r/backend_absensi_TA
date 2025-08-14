import asyncio
from datetime import datetime
import json
import os
from typing import Optional
import uuid
import aiomysql
from fastapi import APIRouter, Query, Depends, File, Form, Request, HTTPException, Security, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import JSONResponse, FileResponse
from koneksi import get_db
from fastapi_jwt import (
  JwtAccessBearerCookie,
  JwtAuthorizationCredentials,
  JwtRefreshBearer
)
import pandas as pd
from aiomysql import Error as aiomysqlerror
from jwt_auth import access_security, refresh_security
import calendar
import time
import hashlib
from utils.fn_conv_str import serialize_data
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from openpyxl.workbook.protection import WorkbookProtection
from collections import defaultdict
from aiomysql import Error as aiomysqlerror
import win32com.client
from pywintypes import com_error
from utils.fn_log import logger

app = APIRouter(
  prefix="/admin"
)

# Ini dari User ke Admin
absensi_connection = []

@app.websocket('/ws-absensi')
async def ws_absen_user(
  websocket: WebSocket
):
  await websocket.accept()
  absensi_connection.append(websocket)

  try:
    print("Hai WS Nyala")
    await websocket.receive_text()
  except WebSocketDisconnect:
    print("WS Disconnect")
    absensi_connection.remove(websocket)
# End Dari user Ke admin


@app.get('/get_absensi')
async def get_absensi(
  request: Request,
  tgl: Optional[str] = Query(None)
):
  try:
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:

        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          kondisi = ""
          params = []
          if tgl:
            kondisi = "WHERE DATE(a.tanggal_absen) = %s"
            params.append(tgl)
          else:
            kondisi = "WHERE DATE(a.tanggal_absen) = DATE(CURDATE())"

          q1 = f"""
            SELECT a.*, k.nama_karyawan, d.nama_departemen FROM absensi a
            INNER JOIN karyawan k ON a.id_karyawan = k.id_karyawan
            INNER JOIN departemen d ON k.id_departemen = d.id_departemen
            {kondisi}
            ORDER BY DATE(a.tanggal_absen) DESC
          """
          await cursor.execute(q1, tuple(params))
          items = await cursor.fetchall()

          return items

        except aiomysqlerror as e:
          return JSONResponse(content={"status": "error", "message": f"Database Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse(content={"status": "error", "message": f"HTTP Error Error {str(e)}"}, status_code=e.status_code)
      
  except Exception as e:
    return JSONResponse(content={"status": "error", "message": f"Koneksi Error {str(e)}"}, status_code=500)
  
@app.get('/get_data_dashboard')
async def get_data_dashboard(
  request: Request
):
  try:
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:

        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          # Absen pending
          q1 = """
            SELECT COUNT(*) as pending FROM absensi WHERE status_absen = "pending" 
          """
          await cursor.execute(q1)
          items = await cursor.fetchone()

          # Total Karyawan
          q2 = """
            SELECT COUNT(*) as karyawan FROM karyawan WHERE status = "aktif" 
          """
          await cursor.execute(q2)
          items2 = await cursor.fetchone()

          # Yang Ga masuk
          q3 = """
            SELECT COUNT(*) as ga_hadir FROM absensi WHERE pengajuan IN ('cuti', 'sakit', 'izin') 
            AND DATE(tanggal_absen) = DATE(CURRENT_TIMESTAMP())
          """
          await cursor.execute(q3)
          items3 = await cursor.fetchone()

          # This is the log message you wanted
          log_message = (
              f"ADMIN MENGAKSES DASHBOARD "
          )
          logger.info(log_message)
          # --- End of logging ---

          return {
            "total_karyawan": items2,
            "absen_pending": items,
            "data_ga_hadir": items3
          }

        except aiomysqlerror as e:
          return JSONResponse(content={"status": "error", "message": f"Database Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse(content={"status": "error", "message": f"HTTP Error Error {str(e)}"}, status_code=e.status_code)

      
  except Exception as e:
    return JSONResponse(content={"status": "error", "message": f"Koneksi Error {str(e)}"}, status_code=500)
  
  
@app.get('/get_pengajuan')
async def get_pengajuan(
  request: Request,
  tgl: Optional[str] = Query(None)
):
  try:
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:

        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")
          kondisi = ""
          params = []
          if tgl:
            kondisi = "WHERE DATE(pa.tanggal_mulai) = %s"
            params.append(tgl)
          else:
            kondisi = "WHERE DATE(pa.tanggal_mulai) = DATE(CURDATE())"

          q1 = f"""
            SELECT pa.*, k.nama_karyawan FROM pengajuan_absen pa
            INNER JOIN karyawan k ON pa.id_karyawan = k.id_karyawan
            {kondisi}
          """
          await cursor.execute(q1, tuple(params))
          items = await cursor.fetchall()

          return items

        except aiomysqlerror as e:
          return JSONResponse(content={"status": "error", "message": f"Database Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse(content={"status": "error", "message": f"HTTP Error Error {str(e)}"}, status_code=e.status_code)

      
  except Exception as e:
    return JSONResponse(content={"status": "error", "message": f"Koneksi Error {str(e)}"}, status_code=500)
  


@app.get('/get_karyawan')
async def get_karyawan(
  request: Request,
  id_karyawan: Optional[str] = Query(None),
):
  try:
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          # Start building the base query
          query = """
              SELECT k.*, d.nama_departemen
              FROM karyawan k
              INNER JOIN departemen d ON k.id_departemen = d.id_departemen
          """
          # Initialize parameters list
          params = []

          # If id_karyawan is provided, add a WHERE clause and corresponding parameter
          if id_karyawan:
            query += " WHERE k.id_karyawan = %s"
            params.append(id_karyawan)

          # Execute the query with the parameters
          await cursor.execute(query, tuple(params))  # Using tuple for params

          items = await cursor.fetchall() if not id_karyawan else await cursor.fetchone()

          return items 

        except aiomysql.MySQLError as e:
            return JSONResponse(content={"status": "error", "message": f"Database Error {str(e)}"}, status_code=500)
        except HTTPException as e:
            return JSONResponse(content={"status": "error", "message": f"HTTP Error Error {str(e)}"}, status_code=e.status_code)

  except Exception as e:
    return JSONResponse(content={"status": "error", "message": f"Koneksi Error {str(e)}"}, status_code=500)


@app.get('/get_exists_akun')
async def get_exists_akun(
  username: str
):
  try:
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:

        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          q1 = """
            SELECT * FROM akun WHERE username = %s
          """
          await cursor.execute(q1, username)
          items = await cursor.fetchall()

          return items if items else []

        except aiomysqlerror as e:
          return JSONResponse(content={"status": "error", "message": f"Database Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse(content={"status": "error", "message": f"HTTP Error Error {str(e)}"}, status_code=e.status_code)

      
  except Exception as e:
    return JSONResponse(content={"status": "error", "message": f"Koneksi Error {str(e)}"}, status_code=500)

@app.get('/get_akun')
async def get_akun(
  request: Request
):
  try:
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:

        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          q1 = """
            SELECT * FROM akun
          """
          await cursor.execute(q1)
          items = await cursor.fetchall()

          return items

        except aiomysqlerror as e:
          return JSONResponse(content={"status": "error", "message": f"Database Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse(content={"status": "error", "message": f"HTTP Error Error {str(e)}"}, status_code=e.status_code)

      
  except Exception as e:
    return JSONResponse(content={"status": "error", "message": f"Koneksi Error {str(e)}"}, status_code=500)
  
@app.get('/get_departemen')
async def get_departemen(
  request: Request
):
  try:
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:

        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          q1 = """
            SELECT * FROM departemen
          """
          await cursor.execute(q1)
          items = await cursor.fetchall()

          return items

        except aiomysqlerror as e:
          return JSONResponse(content={"status": "error", "message": f"Database Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse(content={"status": "error", "message": f"HTTP Error Error {str(e)}"}, status_code=e.status_code)

      
  except Exception as e:
    return JSONResponse(content={"status": "error", "message": f"Koneksi Error {str(e)}"}, status_code=500)
  
@app.get('/get_konfigurasi')
async def get_konfigurasi(
  request: Request
):
  try:
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:

        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          q1 = """
            SELECT * FROM konfigurasi_aplikasi
          """
          await cursor.execute(q1)
          items = await cursor.fetchone()

          return items

        except aiomysqlerror as e:
          return JSONResponse(content={"status": "error", "message": f"Database Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse(content={"status": "error", "message": f"HTTP Error Error {str(e)}"}, status_code=e.status_code)
      
  except Exception as e:
    return JSONResponse(content={"status": "error", "message": f"Koneksi Error {str(e)}"}, status_code=500)
  
@app.get('/get_jadwal')
async def get_jadwal(
  request: Request
):
  try:
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:

        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          q1 = """
            SELECT * FROM jadwal_kerja
          """
          await cursor.execute(q1)
          items = await cursor.fetchall()

          return items

        except aiomysqlerror as e:
          return JSONResponse(content={"status": "error", "message": f"Database Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse(content={"status": "error", "message": f"HTTP Error Error {str(e)}"}, status_code=e.status_code)
      
  except Exception as e:
    return JSONResponse(content={"status": "error", "message": f"Koneksi Error {str(e)}"}, status_code=500)
  
def excel_to_pdf(excel_path, pdf_path):
  excel = win32com.client.Dispatch("Excel.Application")
  excel.Visible = False #Buat Excel Hidden
  excel.DisplayAlerts = False #Lewati Alert

  try:
    print(f"Converting '{excel_path}' to PDF...")
    wb = excel.Workbooks.Open(os.path.abspath(excel_path))
    # --- Add Page Setup to fit content ---
    ws = wb.ActiveSheet
    # --- Page Setup Configuration ---
    # 1. Set print area to only used cells
    used_range = ws.UsedRange
    ws.PageSetup.PrintArea = used_range.Address
    
    # 2. Fit to one page wide and tall
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = 1
    
    # 3. Prevent row/column splitting
    ws.PageSetup.FitToPagesTall = False  # Allow multiple pages if needed
    ws.PageSetup.Zoom = False  # Disable zoom to enforce FitToPages
    
    # 4. Set margins (optional, adjust as needed)
    ws.PageSetup.LeftMargin = 20
    ws.PageSetup.RightMargin = 20
    ws.PageSetup.TopMargin = 20
    ws.PageSetup.BottomMargin = 20
    
    # 5. Center on page
    ws.PageSetup.CenterHorizontally = True
    ws.PageSetup.CenterVertically = True
    
    # 6. Set paper size (A4)
    ws.PageSetup.PaperSize = 9  # 9 = xlPaperA4
    
    # 7. Set orientation (Auto: Excel will decide based on content)
    if used_range.Columns.Count > 10:  # If many columns, use landscape
        ws.PageSetup.Orientation = 2  # 2 = xlLandscape
    else:
        ws.PageSetup.Orientation = 1  # 1 = xlPortrait

    # --- Export to PDF ---
    wb.ActiveSheet.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
    print("Conversion successful!")
        
  except com_error as e:
    print(f"Conversion failed: {e}")
  finally:
    if 'wb' in locals() and wb:
      wb.Close(SaveChanges=False)
    if excel:
      excel.Quit()
  
def formatStrDate(
  params: str
):
  tgl = params.split("-")
  formatted_tgl = tgl[2] + "-" + tgl[1] + "-" + tgl[0]
  return formatted_tgl

# --- Helper Function for Indonesian Date Formatting ---
def format_indonesian_date(date_obj):
    """Formats a date object into 'Nama Hari, DD Nama Bulan YYYY' in Indonesian."""
    days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    months = [
      "Januari", "Februari", "Maret", "April", "Mei", "Juni",
      "Juli", "Agustus", "September", "Oktober", "November", "Desember"
    ]
    day_name = days[date_obj.weekday()]
    month_name = months[date_obj.month - 1]
    return f"{day_name}, {date_obj.day} {month_name} {date_obj.year}"

def bulan_indo(month):
  bulan = ""
  if month == "01":
    bulan = "Januari"
  elif month == "02":
    bulan = "Februari"
  elif month == "03":
    bulan = "Maret"
  elif month == "04":
    bulan = "April"
  elif month == "05":
    bulan = "Mei"
  elif month == "06":
    bulan = "Juni"
  elif month == "07":
    bulan = "Juli"
  elif month == "08":
    bulan = "Agustus"
  elif month == "09":
    bulan = "September"
  elif month == "10":
    bulan = "Oktober"
  elif month == "11":
    bulan = "November"
  elif month == "12":
    bulan = "Desember"

  return bulan


@app.get('/export_excel')
async def exportExcel(
  start_date: Optional[str] = Query(None),
  end_date: Optional[str] = Query(None)
):
  log_message = (
    f"PROSES GENERATE EXCEL REKAPITULASI"
  )
  logger.info(log_message)

  try:
    pool = await get_db()
    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          # --- 1. Determine Date Range and SQL Condition ---
          kondisi = ""
          params = []
          periode_laporan = ""

          if start_date and end_date:
            # Sementara ga pake dlu yg kondisi if ini
            # periode_laporan = f"{formatStrDate(start_date)} s/d {formatStrDate(end_date)}"
            kondisi = "WHERE DATE(a.tanggal_absen) BETWEEN %s AND %s"
            params.extend([start_date, end_date])
          elif start_date:
            tgl = str(start_date).split("-")
            periode_laporan = (bulan_indo(tgl[1]) + " Tahun " + str(tgl[0])).upper()

            kondisi = "WHERE MONTH(a.tanggal_absen) = MONTH(%s)"
            params.append(start_date)
          else: # Default to current month
            await cursor.execute("SELECT CONCAT(YEAR(CURDATE()), '-', LPAD(MONTH(CURDATE()), 2, '0'), '-01') as first_day, LAST_DAY(CURDATE()) as last_day, MONTHNAME(CURDATE()) as month, YEAR(CURDATE()) as year")
            date_info = await cursor.fetchone()
            periode_laporan = f"Bulan {date_info['month']} {date_info['year']}"
            kondisi = "WHERE MONTH(a.tanggal_absen) = MONTH(CURDATE()) AND YEAR(a.tanggal_absen) = YEAR(CURDATE())"
        
          # --- 2. Modify SQL to fetch the date for grouping ---
          # Added 'a.tanggal_absen' for grouping and ordered by date
          q1 = f"""
            SELECT 
              a.tanggal_absen, 
              k.nama_karyawan, 
              k.posisi, 
              TIME(a.check_in) AS absen_masuk, 
              TIME(a.check_out) AS absen_keluar, 
              a.pengajuan, 
              a.is_telat, 
              a.status_absen, 
              a.alasan_penolakan
            FROM absensi a 
            INNER JOIN karyawan k ON a.id_karyawan = k.id_karyawan
            {kondisi}
            ORDER BY a.tanggal_absen ASC, k.nama_karyawan ASC;
          """
          await cursor.execute(q1, params)
          all_data = await cursor.fetchall()

          if not all_data:
            return JSONResponse(
              {"message": f"Tidak ada data absensi untuk periode {periode_laporan}"},
              status_code=404
            )

          # --- 3. Group Data by Date using a Dictionary ---
          grouped_data = defaultdict(list)
          for row in all_data:
            # Use the date part of 'tanggal_absen' as the key
            grouped_data[row['tanggal_absen'].date()].append(row)

          # --- 4. Setup Excel Workbook and Main Headers ---
          wb = Workbook()
          ws = wb.active
          ws.title = "Laporan Absensi"

          # Main Title
          ws.merge_cells('A1:I1')
          corp_cell = ws['A1']
          corp_cell.value = "CV BENGKEL TEKNOLOGI DISTRIBUSI"
          corp_cell.alignment = Alignment(horizontal='center', vertical='center')
          corp_cell.font = Font(bold=True, size=16)

          # Sub Title
          ws.merge_cells('A2:I2')
          ket_cell = ws['A2']
          ket_cell.value = f"LAPORAN ABSENSI PERIODE {periode_laporan}"
          ket_cell.alignment = Alignment(horizontal='center', vertical='center')
          ket_cell.font = Font(bold=True, size=14)
          
          # Add a blank row before data starts
          ws.append([""]) 

          # --- 5. Define Column Headers for the Tables ---
          column_headers = [
            "No", "Nama Karyawan", "Posisi", "Absen Masuk", "Absen Keluar",
            "Pengajuan", "Terlambat", "Status Absen", "Alasan Penolakan"
          ]
          
          # --- 6. Iterate Through Grouped Data and Write to Sheet ---
          # Sort dictionary by date to ensure chronological order
          sorted_dates = sorted(grouped_data.keys())

          for date_key in sorted_dates:
            records_for_the_day = grouped_data[date_key]
            
            # a. Add the formatted date header for the group
            current_row = ws.max_row + 1 # Add space from previous block
            ws.merge_cells(f'A{current_row}:I{current_row}')
            date_header_cell = ws[f'A{current_row}']
            date_header_cell.value = format_indonesian_date(date_key)
            date_header_cell.font = Font(bold=True, size=12)
            date_header_cell.alignment = Alignment(horizontal='left')
            
            # b. Add the table headers for this group
            ws.append(column_headers)
            header_row = ws[ws.max_row]
            for cell in header_row:
              cell.font = Font(bold=True)
              cell.alignment = Alignment(horizontal="center", vertical="center")
              cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
              cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            # c. Add the data rows for this group
            for i, record in enumerate(records_for_the_day, 1):
              # Convert is_telat to a more readable format
              is_telat_str = "Ya" if record.get('is_telat') == 1 else "Tidak"
              
              # Prepare row data, ensuring None/empty values are handled
              row_data = [
                i,
                record.get('nama_karyawan') or "-",
                record.get('posisi') or "-",
                record.get('absen_masuk') or "-",
                record.get('absen_keluar') or "-",
                record.get('pengajuan') or "-",
                is_telat_str,
                record.get('status_absen') or "-",
                record.get('alasan_penolakan') or "-",
              ]
              ws.append(row_data)

            # d. Add a blank row for spacing after each group
            # --- THIS IS THE FIX ---
            # Instead of ws.append([]), append a list with an empty cell.
            ws.append([""]) # ✅ CORRECTED LINE

          # --- 7. Auto Adjust Column Widths ---
          for col_idx, column in enumerate(ws.columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in column:
              if isinstance(cell, MergedCell):
                  continue
              try:
                  if len(str(cell.value)) > max_length:
                      max_length = len(str(cell.value))
              except:
                  pass
              
            # Set a minimum width and add buffer
            adjusted_width = (max_length + 2) * 1.2
            # Special handling for 'No' column
            if col_idx == 1:
              adjusted_width = 5
            ws.column_dimensions[column_letter].width = adjusted_width

          # --- 8. Save and Return the File ---
          file_path = "data_absensi_harian.xlsx"
          if os.path.exists(file_path):
            os.chmod(file_path, 0o644) # Make writable to overwrite
          
          wb.save(file_path)
          os.chmod(file_path, 0o444) # Set back to read-only

          log_message = (
            f"SELESAI GENERATE EXCEL REKAPITULASI"
          )
          logger.info(log_message)

          return FileResponse(
            os.path.abspath(file_path),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename="laporan_absensi_harian.xlsx"
          )

        except aiomysql.Error as e:
          return JSONResponse({"Error": f"Database Error: {str(e)}"}, status_code=500)
        except Exception as e:
          return JSONResponse({"Error": f"An unexpected error occurred: {str(e)}"}, status_code=500)

  except Exception as e:
    return JSONResponse({"Error": f"Failed to connect to the database: {str(e)}"}, status_code=500)